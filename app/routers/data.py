from __future__ import annotations

from datetime import date, datetime, timezone, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from sqlalchemy import func, text
from sqlalchemy.orm import Session, selectinload

try:
    import openpyxl  # type: ignore[import-not-found]
    from openpyxl.styles import Font  # type: ignore[import-not-found]
except ImportError:  # pragma: no cover
    openpyxl = None
    Font = None

from app.database import get_db
from app.auth import get_current_active_user
from app import models
from app.permissions import ensure_permission
from app.utils.tenant import get_tenant_user_ids


router = APIRouter(prefix="/data", tags=["data"])


TENANT_OPERATIONAL_MODELS = (
    ("credit_transactions", models.CreditTransaction),
    ("sale_returns", models.SaleReturn),
    ("purchase_returns", models.PurchaseReturn),
    ("supplier_payments", models.SupplierPayment),
    ("purchases", models.Purchase),
    ("sales", models.Sale),
    ("stock_movements", models.StockMovement),
    ("creditors", models.Creditor),
    ("suppliers", models.Supplier),
    ("products", models.Product),
)

FULL_RESET_TABLE_NAMES = (
    models.PasswordResetToken.__tablename__,
    models.CreditTransaction.__tablename__,
    models.SaleReturn.__tablename__,
    models.PurchaseReturn.__tablename__,
    models.SupplierPayment.__tablename__,
    models.Purchase.__tablename__,
    models.Sale.__tablename__,
    models.StockMovement.__tablename__,
    models.Creditor.__tablename__,
    models.Supplier.__tablename__,
    models.Product.__tablename__,
    models.SystemSettings.__tablename__,
    models.User.__tablename__,
    models.Branch.__tablename__,
)


def _require_admin(current_user: models.User) -> None:
    ensure_permission(current_user, "manage_data", "Admin access required")


def _clear_tenant_operational_data(db: Session, tenant_user_ids: list[int]) -> dict[str, int]:
    deleted_counts: dict[str, int] = {}

    for label, model in TENANT_OPERATIONAL_MODELS:
        deleted_counts[label] = int(
            db.query(model).filter(model.user_id.in_(tenant_user_ids)).delete(
                synchronize_session=False
            )
            or 0
        )

    db.flush()
    return deleted_counts


def _reset_application_database(db: Session) -> list[str]:
    quoted_table_names = ", ".join(f'"{table_name}"' for table_name in FULL_RESET_TABLE_NAMES)
    db.execute(text(f"TRUNCATE TABLE {quoted_table_names} RESTART IDENTITY CASCADE"))
    db.flush()
    return list(FULL_RESET_TABLE_NAMES)


def _serialize_dt(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        # Use ISO format; keep timezone if present
        return value.isoformat()
    # Dates from SQLAlchemy Date columns may be `datetime.date`
    try:
        return value.isoformat()
    except Exception:
        return str(value)


def _serialize_decimal(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, Decimal):
        # Preserve exact value as string
        return str(value)
    return value


def _serialize_num(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, Decimal):
        # Excel handles floats; this is mainly for human consumption.
        try:
            return float(value)
        except Exception:
            return str(value)
    return value


@router.get("/export")
def export_data(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Export tenant data (admin-only) as a JSON file."""
    _require_admin(current_user)

    tenant_user_ids = get_tenant_user_ids(current_user, db)

    branches = (
        db.query(models.Branch)
        .filter(models.Branch.owner_user_id == current_user.id)
        .order_by(models.Branch.id.asc())
        .all()
    )

    products = (
        db.query(models.Product)
        .options(
            selectinload(models.Product.variants),
            selectinload(models.Product.unit_conversions),
        )
        .filter(models.Product.user_id.in_(tenant_user_ids))
        .order_by(models.Product.id.asc())
        .all()
    )

    movements = (
        db.query(models.StockMovement)
        .filter(models.StockMovement.user_id.in_(tenant_user_ids))
        .order_by(models.StockMovement.id.asc())
        .all()
    )

    sales = (
        db.query(models.Sale)
        .filter(models.Sale.user_id.in_(tenant_user_ids))
        .order_by(models.Sale.id.asc())
        .all()
    )

    creditors = (
        db.query(models.Creditor)
        .filter(models.Creditor.user_id.in_(tenant_user_ids))
        .order_by(models.Creditor.id.asc())
        .all()
    )

    credit_transactions = (
        db.query(models.CreditTransaction)
        .filter(models.CreditTransaction.user_id.in_(tenant_user_ids))
        .order_by(models.CreditTransaction.id.asc())
        .all()
    )

    payload: dict[str, Any] = {
        "export_version": 1,
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "tenant": {
            "admin_email": current_user.email,
            "admin_name": current_user.name,
            "business_name": current_user.business_name,
        },
        "branches": [
            {
                "id": b.id,
                "name": b.name,
                "is_active": b.is_active,
                "created_at": _serialize_dt(b.created_at),
            }
            for b in branches
        ],
        "products": [
            {
                "id": p.id,
                "branch_id": p.branch_id,
                "sku": p.sku,
                "barcode": p.barcode,
                "name": p.name,
                "description": p.description,
                "unit": p.unit,
                "measurement_type": getattr(p, "measurement_type", "count"),
                "allows_fractional_sales": getattr(p, "allows_fractional_sales", False),
                "quantity_step": _serialize_decimal(getattr(p, "quantity_step", None)),
                "variant_group": getattr(p, "variant_group", None),
                "variant_label": getattr(p, "variant_label", None),
                "brand": getattr(p, "brand", None),
                "size": getattr(p, "size", None),
                "color": getattr(p, "color", None),
                "shade": getattr(p, "shade", None),
                "pack_size": p.pack_size,
                "category": p.category,
                "supplier": p.supplier,
                "expiry_date": _serialize_dt(p.expiry_date),
                "cost_price": _serialize_decimal(p.cost_price),
                "pack_cost_price": _serialize_decimal(p.pack_cost_price),
                "selling_price": _serialize_decimal(p.selling_price),
                "pack_selling_price": _serialize_decimal(p.pack_selling_price),
                "variants": [
                    {
                        "id": v.id,
                        "label": v.label,
                        "attributes_json": v.attributes_json or {},
                        "is_active": v.is_active,
                        "sort_order": v.sort_order,
                        "created_at": _serialize_dt(v.created_at),
                        "updated_at": _serialize_dt(v.updated_at),
                    }
                    for v in p.variants
                ],
                "unit_conversions": [
                    {
                        "id": conversion.id,
                        "unit_name": conversion.unit_name,
                        "base_quantity": _serialize_decimal(conversion.base_quantity),
                        "is_sale_unit": conversion.is_sale_unit,
                        "is_purchase_unit": conversion.is_purchase_unit,
                        "sort_order": conversion.sort_order,
                        "created_at": _serialize_dt(conversion.created_at),
                        "updated_at": _serialize_dt(conversion.updated_at),
                    }
                    for conversion in p.unit_conversions
                ],
                "created_at": _serialize_dt(p.created_at),
                "updated_at": _serialize_dt(p.updated_at),
            }
            for p in products
        ],
        "stock_movements": [
            {
                "id": m.id,
                "branch_id": m.branch_id,
                "product_id": m.product_id,
                "variant_id": m.variant_id,
                "sale_id": m.sale_id,
                "change": _serialize_decimal(m.change),
                "reason": m.reason,
                "batch_number": m.batch_number,
                "expiry_date": _serialize_dt(m.expiry_date),
                "unit_cost_price": _serialize_decimal(getattr(m, "unit_cost_price", None)),
                "unit_selling_price": _serialize_decimal(getattr(m, "unit_selling_price", None)),
                "location": m.location,
                "created_at": _serialize_dt(m.created_at),
            }
            for m in movements
        ],
        "sales": [
            {
                "id": s.id,
                "branch_id": s.branch_id,
                "product_id": s.product_id,
                "variant_id": s.variant_id,
                "quantity": _serialize_decimal(s.quantity),
                "sale_unit_type": getattr(s, "sale_unit_type", None),
                "pack_quantity": getattr(s, "pack_quantity", None),
                "unit_price": _serialize_decimal(s.unit_price),
                "total_price": _serialize_decimal(s.total_price),
                "customer_name": s.customer_name,
                "payment_method": s.payment_method,
                "amount_paid": _serialize_decimal(s.amount_paid),
                "partial_payment_method": getattr(s, "partial_payment_method", None),
                "client_sale_id": getattr(s, "client_sale_id", None),
                "notes": s.notes,
                "created_at": _serialize_dt(s.created_at),
            }
            for s in sales
        ],
        "creditors": [
            {
                "id": c.id,
                "branch_id": c.branch_id,
                "name": c.name,
                "phone": c.phone,
                "email": c.email,
                "birthday": _serialize_dt(c.birthday),
                "total_debt": _serialize_decimal(c.total_debt),
                "notes": c.notes,
                "created_at": _serialize_dt(c.created_at),
                "updated_at": _serialize_dt(c.updated_at),
            }
            for c in creditors
        ],
        "credit_transactions": [
            {
                "id": ct.id,
                "branch_id": ct.branch_id,
                "creditor_id": ct.creditor_id,
                "sale_id": ct.sale_id,
                "amount": _serialize_decimal(ct.amount),
                "transaction_type": ct.transaction_type,
                "notes": ct.notes,
                "created_at": _serialize_dt(ct.created_at),
            }
            for ct in credit_transactions
        ],
    }

    import json

    filename = f"gel-invent-export-{datetime.now().strftime('%Y-%m-%d')}.json"
    content = json.dumps(payload, ensure_ascii=False).encode("utf-8")

    return Response(
        content=content,
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/xlsx")
def export_data_xlsx(
    days: int = Query(30, ge=1, le=3650, description="How many days back to include"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Export recent tenant data as an Excel workbook (admin-only)."""
    _require_admin(current_user)

    if openpyxl is None:
        raise HTTPException(status_code=500, detail="Excel export is not available")

    tenant_user_ids = get_tenant_user_ids(current_user, db)
    cutoff = datetime.now(timezone.utc) - timedelta(days=days)

    branches = (
        db.query(models.Branch)
        .filter(models.Branch.owner_user_id == current_user.id)
        .order_by(models.Branch.id.asc())
        .all()
    )
    branch_name_by_id = {b.id: b.name for b in branches}

    products = (
        db.query(models.Product)
        .options(
            selectinload(models.Product.variants),
            selectinload(models.Product.unit_conversions),
        )
        .filter(models.Product.user_id.in_(tenant_user_ids))
        .filter(models.Product.updated_at >= cutoff)
        .order_by(models.Product.id.asc())
        .all()
    )

    movements = (
        db.query(models.StockMovement)
        .filter(models.StockMovement.user_id.in_(tenant_user_ids))
        .filter(models.StockMovement.created_at >= cutoff)
        .order_by(models.StockMovement.id.asc())
        .all()
    )

    sales = (
        db.query(models.Sale)
        .filter(models.Sale.user_id.in_(tenant_user_ids))
        .filter(models.Sale.created_at >= cutoff)
        .order_by(models.Sale.id.asc())
        .all()
    )

    product_ids = {p.id for p in products}
    for s in sales:
        product_ids.add(s.product_id)
    for m in movements:
        product_ids.add(m.product_id)

    products_by_id: dict[int, models.Product] = {}
    if product_ids:
        for p in (
            db.query(models.Product)
            .options(
                selectinload(models.Product.variants),
                selectinload(models.Product.unit_conversions),
            )
            .filter(models.Product.user_id.in_(tenant_user_ids))
            .filter(models.Product.id.in_(list(product_ids)))
            .all()
        ):
            products_by_id[p.id] = p

    wb = openpyxl.Workbook()
    ws_products = wb.active
    ws_products.title = "Products"

    ws_sales = wb.create_sheet("Sales")
    ws_movements = wb.create_sheet("Inventory Movements")

    bold = Font(bold=True) if Font else None

    # Products sheet
    prod_headers = [
        "Product ID",
        "Branch",
        "SKU",
        "Barcode",
        "Name",
        "Category",
        "Supplier",
        "Unit",
        "Measurement Type",
        "Allows Fractional Sales",
        "Quantity Step",
        "Product Family",
        "Variant",
        "Brand",
        "Size",
        "Color",
        "Shade",
        "Variant Options",
        "Sale Units",
        "Pack Size",
        "Cost Price",
        "Selling Price",
        "Created At",
        "Updated At",
    ]
    ws_products.append(prod_headers)
    if bold:
        for cell in ws_products[1]:
            cell.font = bold

    for p in products:
        ws_products.append(
            [
                p.id,
                branch_name_by_id.get(p.branch_id) if p.branch_id else None,
                p.sku,
                p.barcode,
                p.name,
                p.category,
                p.supplier,
                p.unit,
                getattr(p, "measurement_type", "count"),
                getattr(p, "allows_fractional_sales", False),
                _serialize_num(getattr(p, "quantity_step", None)),
                getattr(p, "variant_group", None),
                getattr(p, "variant_label", None),
                getattr(p, "brand", None),
                getattr(p, "size", None),
                getattr(p, "color", None),
                getattr(p, "shade", None),
                ", ".join(v.label for v in p.variants if v.label) or None,
                ", ".join(
                    f"{conversion.unit_name} ({_serialize_decimal(conversion.base_quantity)} {p.unit})"
                    for conversion in p.unit_conversions
                    if conversion.is_sale_unit
                ) or None,
                p.pack_size,
                _serialize_num(p.cost_price),
                _serialize_num(p.selling_price),
                _serialize_dt(p.created_at),
                _serialize_dt(p.updated_at),
            ]
        )

    # Sales sheet
    sales_headers = [
        "Sale ID",
        "Branch",
        "Date",
        "Product ID",
        "SKU",
        "Product Name",
        "Variant ID",
        "Sale Unit",
        "Pack Quantity",
        "Quantity",
        "Unit Price",
        "Total Price",
        "Customer",
        "Payment Method",
        "Amount Paid",
        "Partial Payment Method",
        "Notes",
    ]
    ws_sales.append(sales_headers)
    if bold:
        for cell in ws_sales[1]:
            cell.font = bold

    for s in sales:
        p = products_by_id.get(s.product_id)
        ws_sales.append(
            [
                s.id,
                branch_name_by_id.get(s.branch_id) if s.branch_id else None,
                _serialize_dt(s.created_at),
                s.product_id,
                getattr(p, "sku", None),
                getattr(p, "name", None),
                s.variant_id,
                getattr(s, "sale_unit_type", None),
                getattr(s, "pack_quantity", None),
                _serialize_num(s.quantity),
                _serialize_num(s.unit_price),
                _serialize_num(s.total_price),
                s.customer_name,
                s.payment_method,
                _serialize_num(s.amount_paid),
                getattr(s, "partial_payment_method", None),
                s.notes,
            ]
        )

    # Inventory Movements sheet
    mov_headers = [
        "Movement ID",
        "Branch",
        "Date",
        "Product ID",
        "SKU",
        "Product Name",
        "Variant ID",
        "Change",
        "Reason",
        "Batch",
        "Expiry Date",
        "Location",
    ]
    ws_movements.append(mov_headers)
    if bold:
        for cell in ws_movements[1]:
            cell.font = bold

    for m in movements:
        p = products_by_id.get(m.product_id)
        ws_movements.append(
            [
                m.id,
                branch_name_by_id.get(m.branch_id) if m.branch_id else None,
                _serialize_dt(m.created_at),
                m.product_id,
                getattr(p, "sku", None),
                getattr(p, "name", None),
                m.variant_id,
                _serialize_num(m.change),
                m.reason,
                m.batch_number,
                _serialize_dt(m.expiry_date),
                m.location,
            ]
        )

    # Basic column sizing
    for ws in (ws_products, ws_sales, ws_movements):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is None:
                    continue
                val = str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 40)

    buf = BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    filename = f"gel-invent-export-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _as_decimal(value: Any) -> Decimal:
    return Decimal(str(value))


def _as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, Decimal)):
        return bool(value)
    normalized = str(value or "").strip().lower()
    return normalized in {"1", "true", "yes", "on"}


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text_value = str(value).strip()
    return text_value or None


def _as_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _normalized_product_signature_text(value: Any) -> str:
    return (_clean_text(value) or "").lower()


def _as_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    text_value = str(value).strip()
    if not text_value:
        return None
    try:
        return date.fromisoformat(text_value)
    except ValueError:
        try:
            return datetime.fromisoformat(text_value.replace("Z", "+00:00")).date()
        except ValueError:
            return None


@router.post("/clear")
def clear_data(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Reset the application database and restart IDs from 1."""
    _require_admin(current_user)

    truncated_tables = _reset_application_database(db)

    db.commit()
    return {
        "message": "Application database reset completed. All data was removed and IDs were restarted.",
        "truncated_tables": truncated_tables,
        "truncated_count": len(truncated_tables),
    }


@router.post("/import")
def import_data(
    payload: dict[str, Any],
    force: bool = Query(False, description="If true, clears tenant data before import"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Import tenant data from a previous export (admin-only)."""
    _require_admin(current_user)

    if not isinstance(payload, dict) or payload.get("export_version") != 1:
        raise HTTPException(status_code=400, detail="Unsupported export format")

    tenant_user_ids = get_tenant_user_ids(current_user, db)

    # Safety: avoid accidental duplication unless explicitly forced.
    has_any_data = (
        db.query(models.Product.id).filter(models.Product.user_id.in_(tenant_user_ids)).first()
        or db.query(models.Sale.id).filter(models.Sale.user_id.in_(tenant_user_ids)).first()
        or db.query(models.StockMovement.id)
        .filter(models.StockMovement.user_id.in_(tenant_user_ids))
        .first()
        or db.query(models.Creditor.id).filter(models.Creditor.user_id.in_(tenant_user_ids)).first()
        or db.query(models.CreditTransaction.id)
        .filter(models.CreditTransaction.user_id.in_(tenant_user_ids))
        .first()
    )

    if has_any_data and not force:
        raise HTTPException(
            status_code=409,
            detail="Existing data found. Clear data first or re-try with force=true.",
        )

    if has_any_data and force:
        # Delete in dependency order. Keep users/branches.
        db.query(models.CreditTransaction).filter(
            models.CreditTransaction.user_id.in_(tenant_user_ids)
        ).delete(synchronize_session=False)
        db.query(models.Sale).filter(models.Sale.user_id.in_(tenant_user_ids)).delete(
            synchronize_session=False
        )
        db.query(models.StockMovement).filter(
            models.StockMovement.user_id.in_(tenant_user_ids)
        ).delete(synchronize_session=False)
        db.query(models.Creditor).filter(models.Creditor.user_id.in_(tenant_user_ids)).delete(
            synchronize_session=False
        )
        db.query(models.Product).filter(models.Product.user_id.in_(tenant_user_ids)).delete(
            synchronize_session=False
        )
        db.flush()

    branch_id_map: dict[int, int] = {}
    product_id_map: dict[int, int] = {}
    variant_id_map: dict[int, int] = {}
    creditor_id_map: dict[int, int] = {}
    sale_id_map: dict[int, int] = {}
    pending_sale_links: list[tuple[models.StockMovement, int]] = []

    # Branches: upsert by name (for this admin).
    for b in payload.get("branches", []) or []:
        try:
            legacy_id = int(b.get("id"))
            name = str(b.get("name") or "").strip()
        except Exception:
            continue
        if not name:
            continue

        existing = (
            db.query(models.Branch)
            .filter(models.Branch.owner_user_id == current_user.id, models.Branch.name == name)
            .first()
        )
        if existing:
            branch_id_map[legacy_id] = existing.id
            continue

        created = models.Branch(owner_user_id=current_user.id, name=name, is_active=bool(b.get("is_active", True)))
        db.add(created)
        db.flush()
        branch_id_map[legacy_id] = created.id

    # Products
    for p in payload.get("products", []) or []:
        try:
            legacy_id = int(p.get("id"))
            name = str(p.get("name") or "").strip()
            sku = str(p.get("sku") or "").strip()
        except Exception:
            continue
        if not name or not sku:
            continue

        legacy_branch_id = p.get("branch_id")
        new_branch_id = None
        if legacy_branch_id is not None:
            try:
                new_branch_id = branch_id_map.get(int(legacy_branch_id))
            except Exception:
                new_branch_id = None

        # If a product with the same name already exists in this branch, reuse it.
        q = db.query(models.Product).filter(models.Product.user_id == current_user.id)
        if new_branch_id is not None:
            q = q.filter(models.Product.branch_id == new_branch_id)
        existing = q.filter(
            func.lower(func.trim(models.Product.name)) == _normalized_product_signature_text(name),
            func.lower(func.trim(func.coalesce(models.Product.variant_group, ""))) == _normalized_product_signature_text(p.get("variant_group")),
            func.lower(func.trim(func.coalesce(models.Product.variant_label, ""))) == _normalized_product_signature_text(p.get("variant_label")),
            func.lower(func.trim(func.coalesce(models.Product.brand, ""))) == _normalized_product_signature_text(p.get("brand")),
            func.lower(func.trim(func.coalesce(models.Product.size, ""))) == _normalized_product_signature_text(p.get("size")),
            func.lower(func.trim(func.coalesce(models.Product.color, ""))) == _normalized_product_signature_text(p.get("color")),
            func.lower(func.trim(func.coalesce(models.Product.shade, ""))) == _normalized_product_signature_text(p.get("shade")),
        ).first()
        if existing:
            product_id_map[legacy_id] = existing.id
            continue

        product = models.Product(
            user_id=current_user.id,
            branch_id=new_branch_id,
            sku=sku,
            barcode=_clean_text(p.get("barcode")),
            name=name,
            description=_clean_text(p.get("description")),
            unit=str(p.get("unit") or "unit"),
            measurement_type=str(p.get("measurement_type") or "count"),
            allows_fractional_sales=_as_bool(p.get("allows_fractional_sales")),
            quantity_step=_as_decimal(p.get("quantity_step")) if p.get("quantity_step") is not None else None,
            variant_group=_clean_text(p.get("variant_group")),
            variant_label=_clean_text(p.get("variant_label")),
            brand=_clean_text(p.get("brand")),
            size=_clean_text(p.get("size")),
            color=_clean_text(p.get("color")),
            shade=_clean_text(p.get("shade")),
            pack_size=p.get("pack_size"),
            category=_clean_text(p.get("category")),
            supplier=_clean_text(p.get("supplier")),
            expiry_date=_as_date(p.get("expiry_date")),
            cost_price=_as_decimal(p.get("cost_price")) if p.get("cost_price") is not None else None,
            pack_cost_price=_as_decimal(p.get("pack_cost_price")) if p.get("pack_cost_price") is not None else None,
            selling_price=_as_decimal(p.get("selling_price")) if p.get("selling_price") is not None else None,
            pack_selling_price=_as_decimal(p.get("pack_selling_price")) if p.get("pack_selling_price") is not None else None,
        )
        db.add(product)
        db.flush()
        product_id_map[legacy_id] = product.id

    # Product extension tables (variants and unit conversions)
    for p in payload.get("products", []) or []:
        legacy_product_id = _as_int(p.get("id"))
        if legacy_product_id is None:
            continue
        new_product_id = product_id_map.get(legacy_product_id)
        if not new_product_id:
            continue

        for v in p.get("variants", []) or []:
            label = _clean_text(v.get("label"))
            if not label:
                continue

            existing_variant = (
                db.query(models.ProductVariant)
                .filter(models.ProductVariant.product_id == new_product_id)
                .filter(func.lower(func.trim(models.ProductVariant.label)) == label.lower())
                .first()
            )
            if existing_variant is None:
                existing_variant = models.ProductVariant(product_id=new_product_id, label=label)
                db.add(existing_variant)
                db.flush()

            existing_variant.label = label
            existing_variant.attributes_json = v.get("attributes_json") if isinstance(v.get("attributes_json"), dict) else {}
            existing_variant.is_active = _as_bool(v.get("is_active", True))
            sort_order = _as_int(v.get("sort_order"))
            existing_variant.sort_order = sort_order if sort_order is not None else 0
            db.add(existing_variant)
            db.flush()

            legacy_variant_id = _as_int(v.get("id"))
            if legacy_variant_id is not None:
                variant_id_map[legacy_variant_id] = existing_variant.id

        for conversion in p.get("unit_conversions", []) or []:
            unit_name = _clean_text(conversion.get("unit_name"))
            base_quantity = conversion.get("base_quantity")
            if not unit_name or base_quantity is None:
                continue

            existing_conversion = (
                db.query(models.ProductUnitConversion)
                .filter(models.ProductUnitConversion.product_id == new_product_id)
                .filter(func.lower(func.trim(models.ProductUnitConversion.unit_name)) == unit_name.lower())
                .first()
            )
            if existing_conversion is None:
                existing_conversion = models.ProductUnitConversion(
                    product_id=new_product_id,
                    unit_name=unit_name,
                    base_quantity=_as_decimal(base_quantity),
                )
                db.add(existing_conversion)
                db.flush()

            existing_conversion.unit_name = unit_name
            existing_conversion.base_quantity = _as_decimal(base_quantity)
            existing_conversion.is_sale_unit = _as_bool(conversion.get("is_sale_unit", True))
            existing_conversion.is_purchase_unit = _as_bool(conversion.get("is_purchase_unit", False))
            sort_order = _as_int(conversion.get("sort_order"))
            existing_conversion.sort_order = sort_order if sort_order is not None else 0
            db.add(existing_conversion)

    # Stock movements
    for m in payload.get("stock_movements", []) or []:
        try:
            legacy_product_id = int(m.get("product_id"))
        except Exception:
            continue
        new_product_id = product_id_map.get(legacy_product_id)
        if not new_product_id:
            continue

        legacy_branch_id = m.get("branch_id")
        new_branch_id = None
        if legacy_branch_id is not None:
            try:
                new_branch_id = branch_id_map.get(int(legacy_branch_id))
            except Exception:
                new_branch_id = None

        legacy_variant_id = _as_int(m.get("variant_id"))
        new_variant_id = variant_id_map.get(legacy_variant_id) if legacy_variant_id is not None else None

        legacy_sale_id = _as_int(m.get("sale_id"))
        new_sale_id = sale_id_map.get(legacy_sale_id) if legacy_sale_id is not None else None

        movement = models.StockMovement(
            user_id=current_user.id,
            branch_id=new_branch_id,
            product_id=new_product_id,
            variant_id=new_variant_id,
            sale_id=new_sale_id,
            change=_as_decimal(m.get("change", 0)),
            reason=str(m.get("reason") or "adjustment"),
            batch_number=m.get("batch_number"),
            expiry_date=_as_date(m.get("expiry_date")),
            unit_cost_price=_as_decimal(m.get("unit_cost_price")) if m.get("unit_cost_price") is not None else None,
            unit_selling_price=_as_decimal(m.get("unit_selling_price")) if m.get("unit_selling_price") is not None else None,
            location=m.get("location") or "Main Store",
        )
        db.add(movement)
        if legacy_sale_id is not None and new_sale_id is None:
            pending_sale_links.append((movement, legacy_sale_id))

    # Sales
    for s in payload.get("sales", []) or []:
        try:
            legacy_product_id = int(s.get("product_id"))
        except Exception:
            continue
        new_product_id = product_id_map.get(legacy_product_id)
        if not new_product_id:
            continue

        legacy_branch_id = s.get("branch_id")
        new_branch_id = None
        if legacy_branch_id is not None:
            try:
                new_branch_id = branch_id_map.get(int(legacy_branch_id))
            except Exception:
                new_branch_id = None

        legacy_variant_id = _as_int(s.get("variant_id"))
        new_variant_id = variant_id_map.get(legacy_variant_id) if legacy_variant_id is not None else None

        pack_quantity = _as_int(s.get("pack_quantity"))

        sale = models.Sale(
            user_id=current_user.id,
            branch_id=new_branch_id,
            product_id=new_product_id,
            variant_id=new_variant_id,
            quantity=_as_decimal(s.get("quantity", 0)),
            sale_unit_type=str(s.get("sale_unit_type") or "piece"),
            pack_quantity=pack_quantity,
            unit_price=_as_decimal(s.get("unit_price", 0)),
            total_price=_as_decimal(s.get("total_price", 0)),
            customer_name=s.get("customer_name"),
            payment_method=str(s.get("payment_method") or "cash"),
            amount_paid=_as_decimal(s.get("amount_paid")) if s.get("amount_paid") is not None else None,
            partial_payment_method=_clean_text(s.get("partial_payment_method")),
            client_sale_id=_clean_text(s.get("client_sale_id")),
            notes=s.get("notes"),
        )
        db.add(sale)
        db.flush()
        # Not all exports will have stable IDs; keep map only if present.
        if s.get("id") is not None:
            try:
                sale_id_map[int(s.get("id"))] = sale.id
            except Exception:
                pass

    for movement, legacy_sale_id in pending_sale_links:
        new_sale_id = sale_id_map.get(legacy_sale_id)
        if new_sale_id is not None:
            movement.sale_id = new_sale_id
            db.add(movement)

    # Creditors
    for c in payload.get("creditors", []) or []:
        try:
            legacy_id = int(c.get("id"))
            name = str(c.get("name") or "").strip()
        except Exception:
            continue
        if not name:
            continue

        legacy_branch_id = c.get("branch_id")
        new_branch_id = None
        if legacy_branch_id is not None:
            try:
                new_branch_id = branch_id_map.get(int(legacy_branch_id))
            except Exception:
                new_branch_id = None

        creditor = models.Creditor(
            user_id=current_user.id,
            branch_id=new_branch_id,
            name=name,
            phone=c.get("phone"),
            email=c.get("email"),
            birthday=_as_date(c.get("birthday")),
            total_debt=_as_decimal(c.get("total_debt")) if c.get("total_debt") is not None else _as_decimal(0),
            notes=c.get("notes"),
        )
        db.add(creditor)
        db.flush()
        creditor_id_map[legacy_id] = creditor.id

    # Credit transactions
    for ct in payload.get("credit_transactions", []) or []:
        legacy_creditor_id = ct.get("creditor_id")
        if legacy_creditor_id is None:
            continue
        try:
            new_creditor_id = creditor_id_map.get(int(legacy_creditor_id))
        except Exception:
            new_creditor_id = None
        if not new_creditor_id:
            continue

        legacy_sale_id = ct.get("sale_id")
        new_sale_id = None
        if legacy_sale_id is not None:
            try:
                new_sale_id = sale_id_map.get(int(legacy_sale_id))
            except Exception:
                new_sale_id = None

        legacy_branch_id = ct.get("branch_id")
        new_branch_id = None
        if legacy_branch_id is not None:
            try:
                new_branch_id = branch_id_map.get(int(legacy_branch_id))
            except Exception:
                new_branch_id = None

        txn = models.CreditTransaction(
            user_id=current_user.id,
            branch_id=new_branch_id,
            creditor_id=new_creditor_id,
            sale_id=new_sale_id,
            amount=_as_decimal(ct.get("amount", 0)),
            transaction_type=str(ct.get("transaction_type") or "debt"),
            notes=ct.get("notes"),
        )
        db.add(txn)

    db.commit()

    return {"message": "Import completed"}
